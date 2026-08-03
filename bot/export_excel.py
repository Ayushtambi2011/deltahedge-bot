"""Daily Excel export -> 'OUTPUT FILE.xlsx' at the project root.
All numbers are scaled to the $1000 account (per-contract P&L x quantity).
Sheets: Trades (full log + entry context), By Strategy (which one works), Open Positions.

Run nightly (added to desk-nightly workflow). Also runnable locally: python3 export_excel.py
"""
import csv
import json
import os
import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

ROOT = os.path.join(os.path.dirname(__file__), "..")
TRADES = os.path.join(ROOT, "data", "paper_trades.csv")
OUT = os.path.join(ROOT, "OUTPUT FILE.xlsx")
ACCOUNT = 1000.0

TRADE_COLS = ["opened", "symbol", "strategy", "expiry", "qty", "status",
              "entry_spot", "entry_iv", "pop", "bias", "pcr", "max_pain",
              "support", "resistance", "net_delta", "net_theta", "net_vega", "rr",
              "pos_credit_$", "pos_max_loss_$", "settle_spot", "pos_pnl_$", "ret_%_1000",
              "closed", "why"]


def _load():
    if not os.path.exists(TRADES):
        return []
    with open(TRADES) as f:
        return list(csv.DictReader(f))


def _num(x, d=0.0):
    try:
        return float(x)
    except (TypeError, ValueError):
        return d


def _rows():
    out = []
    for r in _load():
        ctx = {}
        try:
            ctx = json.loads(r.get("context") or "{}")
        except json.JSONDecodeError:
            pass
        qty = int(_num(r.get("qty"), 1)) or 1
        pnl = _num(r.get("pnl_net")) * qty if r.get("status") == "settled" else None
        row = {
            "opened": (r.get("opened") or "")[:16], "symbol": r.get("symbol"),
            "strategy": r.get("strategy"), "expiry": r.get("expiry"), "qty": qty,
            "status": r.get("status"),
            "entry_spot": ctx.get("entry_spot"), "entry_iv": ctx.get("entry_iv"),
            "pop": ctx.get("pop"), "bias": ctx.get("bias"), "pcr": ctx.get("pcr"),
            "max_pain": ctx.get("max_pain"), "support": ctx.get("support"),
            "resistance": ctx.get("resistance"), "net_delta": ctx.get("net_delta"),
            "net_theta": ctx.get("net_theta"), "net_vega": ctx.get("net_vega"),
            "rr": ctx.get("rr"),
            "pos_credit_$": round(_num(r.get("net_credit")) * qty, 2),
            "pos_max_loss_$": round(_num(r.get("max_loss")) * qty, 2),
            "settle_spot": r.get("settle_spot"),
            "pos_pnl_$": round(pnl, 2) if pnl is not None else None,
            "ret_%_1000": round(100 * pnl / ACCOUNT, 3) if pnl is not None else None,
            "closed": r.get("closed"), "why": ctx.get("why"),
        }
        out.append(row)
    return out


def _by_strategy(rows):
    g = {}
    for r in rows:
        if r["status"] != "settled" or r["pos_pnl_$"] is None:
            continue
        s = g.setdefault(r["strategy"], {"n": 0, "w": 0, "gw": 0.0, "gl": 0.0})
        s["n"] += 1
        p = r["pos_pnl_$"]
        if p > 0:
            s["w"] += 1; s["gw"] += p
        else:
            s["gl"] += p
    summary = []
    for name, s in sorted(g.items(), key=lambda kv: -(kv[1]["gw"] + kv[1]["gl"])):
        net = s["gw"] + s["gl"]
        summary.append({
            "strategy": name, "settled": s["n"],
            "win_%": round(100 * s["w"] / s["n"], 1) if s["n"] else 0,
            "profit_factor": round(s["gw"] / abs(s["gl"]), 2) if s["gl"] else None,
            "net_$": round(net, 2), "avg_$/trade": round(net / s["n"], 2) if s["n"] else 0,
            "return_%_on_1000": round(100 * net / ACCOUNT, 2),
        })
    return summary


def build():
    if openpyxl is None:
        print("openpyxl not installed — run: pip install openpyxl"); return
    rows = _rows()
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F2A44")
    hdr_font = Font(color="FFFFFF", bold=True)

    def sheet(ws, cols, data):
        ws.append(cols)
        for c in ws[1]:
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")
        for d in data:
            ws.append([d.get(c) for c in cols])
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(10, len(col) + 2)
        ws.freeze_panes = "A2"

    ws1 = wb.active; ws1.title = "Trades"
    sheet(ws1, TRADE_COLS, rows)

    ws2 = wb.create_sheet("By Strategy")
    summ = _by_strategy(rows)
    sheet(ws2, ["strategy", "settled", "win_%", "profit_factor", "net_$",
                "avg_$/trade", "return_%_on_1000"], summ)

    ws3 = wb.create_sheet("Open Positions")
    sheet(ws3, TRADE_COLS, [r for r in rows if r["status"] == "open"])

    note = wb.create_sheet("README")
    note["A1"] = f"Generated {datetime.datetime.utcnow().isoformat()} UTC · PAPER trades · $1000 account"
    note["A2"] = "All $ figures = per-contract P&L x quantity (position-level on $1000)."
    note["A3"] = "By Strategy = which strategy is actually working after fees. Needs settled trades."

    wb.save(OUT)
    print(f"wrote {OUT} · {len(rows)} trades · {len(summ)} strategies scored")


if __name__ == "__main__":
    build()
